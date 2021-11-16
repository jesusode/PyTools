#Read excel with openpyxl

import pprint

from openpyxl import *
import pylightxl as xl

'''
wb = load_workbook("catalogo_suministros_2021.xlsx")
ws = wb.active
print(wb.sheetnames)
rows=[]
for sheet in wb:
    print(sheet.title)
cont=0
for row in ws.iter_rows():
    if cont == 50: break
    rw=[]
    for cell in row:
        rw.append(cell.value)
    rows.append(rw)
    cont +=1


pprint.pprint(rows)
'''

#Con pylightxl
db = xl.readxl(fn='catalogo_suministros_2021.xlsx')
print(db.ws_names)
for row in db.ws(ws=db.ws_names[0]).rows:
    print(row[1])



